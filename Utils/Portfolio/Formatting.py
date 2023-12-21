import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas import DataFrame, Series

## create dictionary for automated number formatting of excel columns by name
mappings_number_formattings = {
    "weight": "0.00%;-0.00%",
    "relative_return": "0.00%;-0.00%",
    "total_return": "0.00%;-0.00%",
    "volatility": "0.00%;-0.00%",
    "beta": "0.00;-0.00",
    "date": "dd.mm.yyyy",
}


## create function to write df to xlsx worksheet, formatted as Table
def write_df_to_xlsx_table(
    wb: workbook,
    ws_name: str,
    df: DataFrame,
    base_formatting: str = "General",
) -> None:
    """Function to write df to an Excel table using openpyxl

    Parameters
    ----------
    wb : openpyxl workbook
        oxl workbook item to which the ws should be added
    ws_name : _type_
        desired worksheet name in Excel output
    df : DataFrame
        df that should be written to Excel
    base_formatting : str, optional
        Excel number formatting applied to all columns not found in mappings_number_formattings, by default "General"
    """
    wb.create_sheet(ws_name)
    ws = wb[ws_name]
    for row in dataframe_to_rows(df.reset_index(), index=False, header=True):
        ws.append(row)
    table = Table(
        displayName=ws_name,
        ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row),
    )

    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)

    ws.add_table(table)

    for column in ws.columns:
        cell_width = max(len(str(cell.value)) for cell in column) * 1.33
        ws.column_dimensions[column[0].column_letter].width = cell_width
    ws.freeze_panes = "B2"

    for column_name in df.reset_index().columns:
        column_letter = get_column_letter(
            df.reset_index().columns.get_loc(column_name) + 1
        )
        if mappings_number_formattings.get(column_name):
            number_formatting = mappings_number_formattings.get(column_name)
        else:
            number_formatting = base_formatting
        for cell in ws[column_letter]:
            cell.number_format = number_formatting


## create function to plot line charts for stock returns / prices
def line_plot(
    series: list[Series] | Series,
    title: str | None = None,
    xlabel: str | None = None,
    ylabel: str | None = None,
    format_as_pct=True,
):
    ### if series input, convert to list of series
    if isinstance(series, Series):
        series = [series]

    plt.figure(figsize=(10, 5))

    ### plot all series
    for s in series:
        s.plot(label=s.name)

    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)

    ### base line chart params
    plt.grid(True)
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()

    if format_as_pct:
        plt.gca().yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1))

    plt.show()
